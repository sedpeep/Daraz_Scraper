import time
import pandas as pd
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException

options = Options()
options.add_experimental_option("detach", True)

driver = webdriver.Edge(options=options)
driver.get("https://www.daraz.pk/?spm=a2a0e.pdp.header.dhome.5647qJlnqJlnm0")

# Get Links
product = input("Enter the product name:")
search = driver.find_element(By.XPATH, "//*[@id='q']")
search.send_keys(product)
#Wait until search button is found
search_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='topActionHeader']/div/div[2]/div/div[2]/form/div/div[2]/button")))
driver.execute_script("arguments[0].scrollIntoView(true);",search_button)
search_button.click()

# Function to extract data from the current page
def extract_data():
    description = driver.find_elements(By.XPATH, "//div[starts-with(@class,'title--wFj93')]//a")
    image = driver.find_elements(By.XPATH, "//div[starts-with(@class,'mainPic--ehOdr')]//img")
    price=driver.find_elements(By.XPATH, "//div[starts-with(@class,'price--NVB62')]")
    discount=driver.find_elements(By.XPATH,".//div[starts-with(@class, 'priceExtra--ocAYk')]//del")
    product_link=driver.find_elements(By.XPATH,".//div[starts-with(@class, 'mainPic--ehOdr')]//a")

    product_list = images_list = price_list = discounted_list = product_link_list= []

    for prod, img,prices,discounts,links in zip(description, image,price,discount,product_link):
        #store data in each list
        product_list.append(prod.text)
        images_list.append(img.get_attribute("src"))
        price_list.append(prices.text)
        discounted_list.append(discounts.text)
        product_link_list.append(links.get_attribute("href"))

        #print the data
        print(prod.text)
        print(img.get_attribute("src"))
        print(prices.text)
        print(discounts.text)
        print(links.get_attribute("href"))


    return product_list, images_list,price_list,discounted_list,product_link_list

# Initial extraction from the first page
all_product_list, all_images_list, all_price_list, all_discounted_list, all_product_link_list= extract_data()

# Loop to navigate through multiple pages
page = 1
while page < 5:
    try:
        # Extract data from the current page
        current_product_list, current_images_list, current_price_list, current_discount_list, current_link_list = extract_data()

        # Add data from the current page to the lists
        all_product_list.extend(current_product_list)
        all_images_list.extend(current_images_list)
        all_price_list.extend(current_price_list)
        all_discounted_list.extend(current_discount_list)
        all_product_link_list.extend(current_link_list)
        time.sleep(3)

        # Find and click on the "Next" button
        next_button = driver.find_element(By.XPATH, "//*[@id='root']/div/div[2]/div/div/div[1]/div[3]/div/ul/li[9]/a")

        # Scroll to the top of the page to avoid element interception
        driver.execute_script("window.scrollTo(0, 0);")

        # Wait for the "Next" button to be clickable
        next_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div[2]/div/div/div[1]/div[3]/div/ul/li[9]/a")))
        next_button.click()
        time.sleep(3)  # Give the page time to load
        page += 1

    except ElementClickInterceptedException:
        # If the "Next" button is still intercepted, scroll to the top of the page again
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)  # Wait for the page to settle
        continue

    except NoSuchElementException:
        # If the "Next" button is not found, exit the loop
        break

data_rows = []
for prod, img, price, discount, link in zip(all_product_list, all_images_list, all_price_list, all_discounted_list, all_product_link_list):
    data_rows.append([prod, img, price, discount, link])

# Create Excel Workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

headers = ['Product Name', 'Image URL', 'Price', 'Discounted Price', 'Product Link']

# Write headers to the first row of the worksheet
for col_index, header in enumerate(headers, start=1):
    worksheet.cell(row=1, column=col_index, value=header)

# Write data to the worksheet
for row_idx, data_row in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(data_row, start=1):
        worksheet.cell(row=row_idx, column=col_idx, value=value)

#save to file
# Example of setting an absolute file path
file_path = r"C:\Users\Omer Habib\PycharmProjects\DarazScraper"
file_name = f"{product}_list_5.xlsx"
full_file_path = os.path.join(file_path, file_name)


# Print a message indicating successful data storage
print(f"Data has been successfully scraped and stored in '{file_name}'.")





# Close the browser
driver.quit()